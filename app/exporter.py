"""
exporter.py — экспорт данных в CSV и Excel.
Отправляет файлы прямо в Telegram.
"""
import csv
import io
from collections import defaultdict
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, select

from app.config import logger
from app.database import Database, Product, PriceHistory

MARKETPLACE_COLORS = {
    "ozon":        "FF0055FF",  # синий
    "wildberries": "FFAA00CC",  # фиолетовый
    "aliexpress":  "FFFF4400",  # оранжевый
}


async def get_export_data(db: Database, days: int = 30) -> list[dict]:
    """Собирает все данные для экспорта."""
    since = datetime.now(timezone.utc) - timedelta(days=days)

    async with db.session() as s:
        products = (await s.execute(select(Product).order_by(Product.last_check.desc()))).scalars().all()
        if not products:
            return []

        product_ids = [product.id for product in products]
        latest_at = (
            select(
                PriceHistory.product_id.label("product_id"),
                func.max(PriceHistory.recorded_at).label("recorded_at"),
            )
            .where(PriceHistory.product_id.in_(product_ids))
            .group_by(PriceHistory.product_id)
            .subquery()
        )
        latest_rows = (await s.execute(
            select(PriceHistory)
            .join(
                latest_at,
                and_(
                    PriceHistory.product_id == latest_at.c.product_id,
                    PriceHistory.recorded_at == latest_at.c.recorded_at,
                ),
            )
        )).scalars().all()
        history_rows = (await s.execute(
            select(PriceHistory)
            .where(PriceHistory.product_id.in_(product_ids))
            .where(PriceHistory.recorded_at >= since)
            .order_by(PriceHistory.product_id, PriceHistory.recorded_at)
        )).scalars().all()

    latest_by_product = {row.product_id: row for row in latest_rows}
    history_by_product: dict[int, list[PriceHistory]] = defaultdict(list)
    for item in history_rows:
        history_by_product[item.product_id].append(item)

    rows = []
    for product in products:
        last = latest_by_product.get(product.id)
        history = history_by_product.get(product.id, [])
        prices = [h.price for h in history if h.price]
        min_price = min(prices) if prices else None
        max_price = max(prices) if prices else None
        first_price = prices[0] if prices else None
        last_price = prices[-1] if prices else None

        # Тренд
        trend = "→"
        if first_price and last_price:
            if last_price < first_price * 0.97:
                trend = "📉"
            elif last_price > first_price * 1.03:
                trend = "📈"

        # Маркетплейс
        marketplace = "ozon" if "ozon" in product.url else \
                      "wildberries" if "wildberries" in product.url else "unknown"
        if "market.yandex" in product.url:
            marketplace = "yandex_market"

        rows.append({
            "ID": product.id,
            "Маркетплейс": marketplace.capitalize(),
            "Название": product.name or "",
            "Текущая цена (₽)": last.price if last else None,
            "Мин цена за 30д (₽)": min_price,
            "Макс цена за 30д (₽)": max_price,
            "Изменение цены": trend,
            "Наличие": last.availability_status if last else "unknown",
            "Изменений цены": len(history),
            "Первая проверка": product.first_seen.strftime("%d.%m.%Y") if product.first_seen else "",
            "Последняя проверка": product.last_check.strftime("%d.%m.%Y %H:%M") if product.last_check else "",
            "URL": product.url,
        })

    return rows


def make_csv(data: list[dict]) -> io.BytesIO:
    """Создаёт CSV файл в памяти."""
    buf = io.StringIO()
    if not data:
        buf.write("Нет данных\n")
    else:
        writer = csv.DictWriter(buf, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    result = io.BytesIO(buf.getvalue().encode("utf-8-sig"))  # utf-8-sig для Excel
    result.name = f"prices_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return result


def make_excel(data: list[dict]) -> io.BytesIO:
    """Создаёт красивый Excel файл с форматированием."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мониторинг цен"

    if not data:
        ws["A1"] = "Нет данных"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    headers = list(data[0].keys())

    # ── Стили ────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    gray_fill  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # ── Заголовок листа ──────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"📊 Мониторинг цен маркетплейсов | {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # ── Шапка таблицы ────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 35

    # ── Данные ───────────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(data, 3):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(key == "Название"))

            # Чередование строк
            if row_idx % 2 == 0:
                cell.fill = gray_fill

            # Цветовая кодировка по наличию
            if key == "Наличие":
                if row_data[key] == "in_stock":
                    cell.fill = green_fill
                    cell.value = "✅ В наличии"
                elif row_data[key] == "out_of_stock":
                    cell.fill = red_fill
                    cell.value = "❌ Нет"
                elif row_data[key] == "deleted":
                    cell.fill = red_fill
                    cell.value = "🗑 Удалён"

            # Цветовая кодировка маркетплейса
            if key == "Маркетплейс":
                mp = str(row_data[key]).lower()
                color = MARKETPLACE_COLORS.get(mp, "FFF0F0F0")
                cell.fill = PatternFill(start_color=color[2:], end_color=color[2:], fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Цены числами
            if "цена" in key.lower() and row_data[key]:
                cell.number_format = '#,##0 "₽"'

        ws.row_dimensions[row_idx].height = 22

    # ── Ширина колонок ───────────────────────────────────────────────────────
    col_widths = {
        "ID": 6, "Маркетплейс": 15, "Название": 45,
        "Текущая цена (₽)": 18, "Мин цена за 30д (₽)": 20, "Макс цена за 30д (₽)": 20,
        "Изменение цены": 16, "Наличие": 16, "Изменений цены": 16,
        "Первая проверка": 16, "Последняя проверка": 20, "URL": 50,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)

    # ── Закрепить строки заголовка ───────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Лист с историей цен ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("История цен")
    ws2.append(["Товар", "Цена (₽)", "Статус", "Дата"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2["C1"].font = Font(bold=True)
    ws2["D1"].font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = f"prices_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buf


async def export_csv(db: Database) -> io.BytesIO:
    data = await get_export_data(db)
    return make_csv(data)


async def export_excel(db: Database) -> io.BytesIO:
    data = await get_export_data(db)
    return make_excel(data)

import json
import re
import time
from dataclasses import dataclass
from html import escape
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote_plus, unquote, urljoin, urlparse
from urllib.request import Request, urlopen
from zipfile import ZipFile
from xml.etree import ElementTree as ET

import openpyxl
from bs4 import BeautifulSoup


BRD_SHEET_NAME = "BRD ТбД"
BRD_REQUIRED_COLUMNS = [
    "Артикул",
    "Наименование",
    "Доп наименование",
    "Описание",
    "Характеристики",
    "Картинки",
    "Категория",
]
BRD_SERVICE_COLUMNS = [
    "ИИ поисковый запрос",
    "ИИ статус",
    "ИИ источники",
    "ИИ уверенность",
    "ИИ примечание",
]
CHECKPOINT_SUFFIX = ".checkpoint.json"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36"
)


@dataclass(slots=True)
class BrdSource:
    url: str
    title: str
    text: str
    image_url: str | None = None


@dataclass(slots=True)
class BrdResearchResult:
    status: str
    name: str = ""
    extra_name: str = ""
    description_html: str = ""
    specs_html: str = ""
    image_filename: str = ""
    category: str = ""
    sources: list[str] = None
    confidence: float = 0.0
    note: str = ""

    def __post_init__(self) -> None:
        self.sources = self.sources or []


def prepare_brd_table(
    input_path: str | Path,
    output_path: str | Path,
    *,
    categories_path: str | Path | None = None,
    limit: int | None = None,
    resume: bool = False,
    online: bool = False,
    img_dir: str | Path | None = None,
    delay_sec: float = 1.0,
) -> dict[str, Any]:
    input_path = Path(input_path)
    output_path = Path(output_path)
    categories_path = Path(categories_path) if categories_path else _auto_find_categories(input_path.parent)
    checkpoint_path = output_path.with_suffix(output_path.suffix + CHECKPOINT_SUFFIX)

    categories = load_docx_lines(categories_path)
    workbook = openpyxl.load_workbook(input_path)
    sheet = workbook[BRD_SHEET_NAME] if BRD_SHEET_NAME in workbook.sheetnames else workbook.active
    header_map = _header_map(sheet)
    _validate_headers(header_map)
    _ensure_service_columns(sheet, header_map)
    header_map = _header_map(sheet)

    processed_ids = _load_checkpoint(checkpoint_path) if resume else set()
    img_dir = Path(img_dir) if img_dir else output_path.parent / "img"
    processed = 0
    skipped = 0
    total_rows = 0

    for row_idx in range(2, sheet.max_row + 1):
        article = str(sheet.cell(row_idx, header_map["Артикул"]).value or "").strip()
        if not article:
            continue
        total_rows += 1
        row_id = f"article:{article.lower()}"
        if row_id in processed_ids:
            skipped += 1
            continue
        if limit is not None and processed >= limit:
            continue

        _prepare_brd_row(
            sheet,
            row_idx,
            header_map,
            categories,
            online=online,
            img_dir=img_dir,
        )
        processed_ids.add(row_id)
        processed += 1
        _save_checkpoint(checkpoint_path, processed_ids)
        if online and delay_sec > 0:
            time.sleep(delay_sec)

    output_path.parent.mkdir(parents=True, exist_ok=True) if output_path.parent != Path("") else None
    workbook.save(output_path)
    return {
        "input": str(input_path),
        "output": str(output_path),
        "categories": str(categories_path),
        "checkpoint": str(checkpoint_path),
        "sheet": sheet.title,
        "categories_count": len(categories),
        "total_rows": total_rows,
        "processed": processed,
        "skipped": skipped,
        "resume": resume,
        "online": online,
        "img_dir": str(img_dir),
    }


def load_docx_lines(path: str | Path) -> list[str]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"DOCX file not found: {path}")
    with ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    lines = []
    for paragraph in root.findall(".//w:p", ns):
        text = "".join(t.text or "" for t in paragraph.findall(".//w:t", ns)).strip()
        if text:
            lines.append(re.sub(r"\s+", " ", text))
    return lines


def brady_search_query(article: str) -> str:
    normalized = normalize_brd_article(article)
    return f"Brady {normalized}" if normalized else f"Brady {article.strip()}"


def normalize_brd_article(article: str) -> str:
    value = str(article or "").strip()
    match = re.search(r"(\d[\dA-Za-z-]*)$", value)
    if match:
        return match.group(1)
    return re.sub(r"^[A-Za-zА-Яа-я]+", "", value).strip()


def classify_brd_category(text: str, categories: list[str]) -> str:
    low = (text or "").lower()
    if not low:
        return ""
    if any(needle in low for needle in ("arc flash", "hazard label", "warning label")):
        category = _find_category(
            categories,
            "\u0413\u043e\u0442\u043e\u0432\u044b\u0435 \u0437\u043d\u0430\u043a\u0438 \u0431\u0435\u0437\u043e\u043f\u0430\u0441\u043d\u043e\u0441\u0442\u0438",
        )
        if category:
            return category

    keyword_rules = [
        (("термотрансфер", "thermal transfer"), "Термотрансферные принтеры"),
        (("портатив", "bmp21", "m210", "m211"), "Портативные принтеры"),
        (("стационар", "bbp", "bradyprinter"), "Стационарные принтеры"),
        (("картридж", "ribbon", "tape"), "Картриджи для портативного принтера BMP21 PLUS, M210, M211"),
        (("знак", "safety sign", "безопас"), "Готовые знаки безопасности"),
        (("lockout", "tagout", "loto", "блокиратор"), "Блокираторы и элементы системы Lockout/Tagout (LOTO)"),
        (("сорбент", "spill"), "Сорбенты для предотвращения и ликвидация проливов"),
        (("software", "программ"), "Программное обеспечение"),
    ]
    for needles, category in keyword_rules:
        if any(needle in low for needle in needles) and category in categories:
            return category

    best = ""
    best_score = 0
    words = set(re.findall(r"[a-zа-яё0-9]{4,}", low, flags=re.I))
    for category in categories:
        category_words = set(re.findall(r"[a-zа-яё0-9]{4,}", category.lower(), flags=re.I))
        score = len(words & category_words)
        if score > best_score:
            best = category
            best_score = score
    return best if best_score >= 2 else ""


def _find_category(categories: list[str], target: str) -> str:
    for category in categories:
        if category == target:
            return category
    return ""


def research_brd_article(
    article: str,
    categories: list[str],
    *,
    img_dir: str | Path,
    max_results: int = 3,
) -> BrdResearchResult:
    normalized = normalize_brd_article(article)
    query = brady_search_query(article)
    search_limit = max(max_results * 3, 10)
    urls = search_web_urls(query, limit=search_limit)
    sources: list[BrdSource] = []
    checked = 0
    for url in urls:
        try:
            source = fetch_brd_source(url, normalized)
        except Exception:
            continue
        checked += 1
        evidence_text = "\n".join([source.title, source.url, source.text])
        if _article_match_count(evidence_text, normalized) >= 2:
            sources.append(source)
        if len(sources) >= max_results:
            break

    if not sources:
        return BrdResearchResult(
            status="not_found",
            confidence=0.0,
            note="Не найдены источники с точным вхождением артикула минимум 2 раза.",
        )

    name = _choose_name(sources, article)
    extra_name = f"Артикул {normalized}" if normalized else ""
    summary = _choose_summary(sources)
    specs = _extract_specs(sources, normalized)
    category = classify_brd_category(" ".join([name, summary, " ".join(specs)]), categories)
    image_filename = _download_first_image(sources, article, img_dir)
    confidence = min(0.95, 0.45 + len(sources) * 0.15 + (0.10 if specs else 0) + (0.10 if image_filename else 0))

    return BrdResearchResult(
        status="filled_online",
        name=name,
        extra_name=extra_name,
        description_html=_description_html(summary, sources),
        specs_html=_specs_html(normalized, specs),
        image_filename=image_filename,
        category=category,
        sources=[source.url for source in sources],
        confidence=round(confidence, 2),
        note=f"Заполнено по {len(sources)} источникам с проверкой артикула.",
    )


def search_web_urls(query: str, limit: int = 3) -> list[str]:
    search_url = f"https://duckduckgo.com/html/?q={quote_plus(query)}"
    html = _http_get_text(search_url)
    soup = BeautifulSoup(html, "lxml")
    urls: list[str] = []
    for link in soup.select("a.result__a, a[href]"):
        href = link.get("href") or ""
        url = _normalize_search_href(href)
        if not url or not url.startswith(("http://", "https://")):
            continue
        host = urlparse(url).netloc.lower()
        if "duckduckgo.com" in host:
            continue
        if url not in urls:
            urls.append(url)
        if len(urls) >= limit:
            break
    return urls


def fetch_brd_source(url: str, article: str) -> BrdSource:
    html = _http_get_text(url)
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    title = _first_text(
        soup.select_one("meta[property='og:title']"),
        attr="content",
    ) or _first_text(soup.select_one("h1")) or _first_text(soup.select_one("title")) or url
    image_url = _first_text(soup.select_one("meta[property='og:image']"), attr="content")
    if image_url:
        image_url = urljoin(url, image_url)
    text = soup.get_text("\n", strip=True)
    text = re.sub(r"\n{2,}", "\n", text)
    return BrdSource(url=url, title=_clean_text(title), text=text[:120_000], image_url=image_url)


def _prepare_brd_row(
    sheet,
    row_idx: int,
    header_map: dict[str, int],
    categories: list[str],
    *,
    online: bool = False,
    img_dir: Path | None = None,
) -> None:
    article = str(sheet.cell(row_idx, header_map["Артикул"]).value or "").strip()
    name = str(sheet.cell(row_idx, header_map["Наименование"]).value or "").strip()
    description = str(sheet.cell(row_idx, header_map["Описание"]).value or "").strip()
    specs = str(sheet.cell(row_idx, header_map["Характеристики"]).value or "").strip()
    current_category = str(sheet.cell(row_idx, header_map["Категория"]).value or "").strip()

    query = brady_search_query(article)
    category = current_category or classify_brd_category(" ".join([name, description, specs]), categories)
    if category and not current_category:
        sheet.cell(row_idx, header_map["Категория"], category)

    if online:
        result = research_brd_article(article, categories, img_dir=img_dir or Path("img"))
        if result.status == "filled_online":
            if not name:
                sheet.cell(row_idx, header_map["Наименование"], result.name)
            if not sheet.cell(row_idx, header_map["Доп наименование"]).value:
                sheet.cell(row_idx, header_map["Доп наименование"], result.extra_name)
            if not description:
                sheet.cell(row_idx, header_map["Описание"], result.description_html)
            if not specs:
                sheet.cell(row_idx, header_map["Характеристики"], result.specs_html)
            if result.image_filename and not sheet.cell(row_idx, header_map["Картинки"]).value:
                sheet.cell(row_idx, header_map["Картинки"], result.image_filename)
            if result.category and not current_category:
                sheet.cell(row_idx, header_map["Категория"], result.category)
        sheet.cell(row_idx, header_map["ИИ поисковый запрос"], query)
        sheet.cell(row_idx, header_map["ИИ статус"], result.status)
        sheet.cell(row_idx, header_map["ИИ источники"], ", ".join(result.sources))
        sheet.cell(row_idx, header_map["ИИ уверенность"], f"{result.confidence:.2f}")
        sheet.cell(row_idx, header_map["ИИ примечание"], result.note)
        return

    if name and description and specs:
        status = "prepared_existing_data"
        confidence = "0.60"
        note = "Есть базовые данные; нужна онлайн-проверка источников и картинки."
    else:
        status = "needs_online_research"
        confidence = "0.00"
        note = (
            "По ТЗ нужно искать первые релевантные источники по артикулу, "
            "проверять точное вхождение артикула минимум 2 раза и заполнять HTML."
        )

    sheet.cell(row_idx, header_map["ИИ поисковый запрос"], query)
    sheet.cell(row_idx, header_map["ИИ статус"], status)
    sheet.cell(row_idx, header_map["ИИ источники"], "")
    sheet.cell(row_idx, header_map["ИИ уверенность"], confidence)
    sheet.cell(row_idx, header_map["ИИ примечание"], note)


def _http_get_text(url: str, timeout: int = 20) -> str:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept-Language": "ru,en;q=0.9"})
    with urlopen(request, timeout=timeout) as response:
        content_type = response.headers.get("Content-Type", "")
        raw = response.read(1_500_000)
    charset_match = re.search(r"charset=([\w-]+)", content_type, flags=re.I)
    encoding = charset_match.group(1) if charset_match else "utf-8"
    return raw.decode(encoding, errors="replace")


def _http_get_bytes(url: str, timeout: int = 20) -> tuple[bytes, str]:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "image/avif,image/webp,image/*,*/*"})
    with urlopen(request, timeout=timeout) as response:
        content_type = response.headers.get("Content-Type", "")
        return response.read(5_000_000), content_type


def _normalize_search_href(href: str) -> str:
    if href.startswith("//duckduckgo.com/l/") or href.startswith("/l/"):
        parsed = urlparse(href if href.startswith("http") else "https://duckduckgo.com" + href)
        uddg = parse_qs(parsed.query).get("uddg", [""])[0]
        return unquote(uddg)
    return href


def _article_match_count(text: str, article: str) -> int:
    if not article:
        return 0
    return len(re.findall(rf"(?<![A-Za-z0-9]){re.escape(article)}(?![A-Za-z0-9])", text, flags=re.I))


def _first_text(node, attr: str | None = None) -> str:
    if not node:
        return ""
    if attr:
        return str(node.get(attr) or "").strip()
    return node.get_text(" ", strip=True)


def _choose_name(sources: list[BrdSource], article: str) -> str:
    normalized = normalize_brd_article(article)
    for source in sources:
        title = _clean_text(source.title)
        if title and normalized.lower() in title.lower():
            return _shorten_title(title)
    return _shorten_title(sources[0].title or f"Brady {normalized}")


def _choose_summary(sources: list[BrdSource]) -> str:
    candidates = []
    for source in sources:
        for line in source.text.splitlines():
            line = _clean_text(line)
            if 80 <= len(line) <= 700 and any(word in line.lower() for word in ("brady", "label", "printer", "марки", "этикет", "лента")):
                candidates.append(line)
    if candidates:
        return candidates[0]
    text = _clean_text(sources[0].text)
    return text[:700]


def _extract_specs(sources: list[BrdSource], article: str) -> list[str]:
    specs = [f"Артикул: {normalize_brd_article(article)}", "Производитель: Brady"]
    seen = {spec.lower() for spec in specs}
    spec_keywords = (
        "material", "color", "colour", "width", "height", "length", "size", "compatible",
        "материал", "цвет", "ширина", "высота", "длина", "размер", "совместим",
    )
    for source in sources:
        for line in source.text.splitlines():
            clean = _clean_text(line)
            if not (8 <= len(clean) <= 180):
                continue
            if not any(keyword in clean.lower() for keyword in spec_keywords):
                continue
            key = clean.lower()
            if key in seen:
                continue
            specs.append(clean)
            seen.add(key)
            if len(specs) >= 12:
                return specs
    return specs


def _description_html(summary: str, sources: list[BrdSource]) -> str:
    parts = [f"<p>{escape(_clean_text(summary))}</p>"]
    if len(sources) > 1:
        parts.append(f"<p>Данные сверены по {len(sources)} источникам.</p>")
    return "".join(parts)


def _specs_html(article: str, specs: list[str]) -> str:
    items = "".join(f"<li>{escape(_clean_text(spec))}</li>" for spec in specs if _clean_text(spec))
    return f"<ul>{items}</ul>" if items else ""


def _download_first_image(sources: list[BrdSource], article: str, img_dir: Path) -> str:
    image_url = next((source.image_url for source in sources if source.image_url), None)
    if not image_url:
        return ""
    try:
        data, content_type = _http_get_bytes(image_url)
    except Exception:
        return ""
    if content_type and not content_type.lower().startswith("image/"):
        return ""
    ext = _image_ext(image_url, content_type)
    filename = f"{str(article).strip().lower()}{ext}"
    img_dir.mkdir(parents=True, exist_ok=True)
    (img_dir / filename).write_bytes(data)
    return filename


def _image_ext(url: str, content_type: str) -> str:
    path_ext = Path(urlparse(url).path).suffix.lower()
    if path_ext in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        return path_ext
    if "png" in content_type:
        return ".png"
    if "webp" in content_type:
        return ".webp"
    return ".jpg"


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def _shorten_title(title: str) -> str:
    title = re.sub(r"\s*[\-|–|—]\s*(Brady|Shop|Buy|Online).*$", "", title, flags=re.I)
    return _clean_text(title)[:250]


def _auto_find_categories(directory: Path) -> Path:
    matches = sorted(directory.glob("*Категории*BRD*.docx"))
    if matches:
        return matches[0]
    matches = sorted(directory.glob("*BRD*.docx"))
    if matches:
        return matches[0]
    raise FileNotFoundError("Categories DOCX not found. Pass --brd-categories path.")


def _header_map(sheet) -> dict[str, int]:
    result = {}
    for col_idx in range(1, sheet.max_column + 1):
        value = sheet.cell(1, col_idx).value
        if value is not None:
            result[str(value).strip()] = col_idx
    return result


def _validate_headers(header_map: dict[str, int]) -> None:
    missing = [column for column in BRD_REQUIRED_COLUMNS if column not in header_map]
    if missing:
        raise ValueError(f"BRD table is missing required columns: {', '.join(missing)}")


def _ensure_service_columns(sheet, header_map: dict[str, int]) -> None:
    next_col = sheet.max_column + 1
    for column in BRD_SERVICE_COLUMNS:
        if column in header_map:
            continue
        sheet.cell(1, next_col, column)
        next_col += 1


def _load_checkpoint(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return set(payload.get("processed_ids", []))
    except Exception:
        return set()


def _save_checkpoint(path: Path, processed_ids: set[str]) -> None:
    path.write_text(
        json.dumps({"processed_ids": sorted(processed_ids)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

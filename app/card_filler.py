import hashlib
import io
import json
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime
from statistics import median
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.ai_client import ai_is_available, ask_ai
from app.config import logger


@dataclass(slots=True)
class OzonCardDraft:
    offer_id: str
    name: str
    description: str
    category_hint: str | None = None
    brand: str = "Нет бренда"
    price: int | None = None
    old_price: int | None = None
    vat: str = "0"
    weight_g: int | None = None
    width_mm: int | None = None
    height_mm: int | None = None
    depth_mm: int | None = None
    images: list[str] = field(default_factory=list)
    attributes: dict[str, str] = field(default_factory=dict)
    keywords: list[str] = field(default_factory=list)
    checklist: list[str] = field(default_factory=list)
    ai_notes: str = ""
    competitor_summary: str = ""
    selling_points: list[str] = field(default_factory=list)
    source_text: str = ""


@dataclass(slots=True)
class OzonCardBatchItem:
    draft: OzonCardDraft | None
    source: str
    status: str
    message: str = ""


def build_ozon_card_draft(task: str) -> OzonCardDraft:
    text = _clean(task)
    urls = _extract_urls(text)
    price = _extract_price(text)
    old_price = round(price * 1.25) if price else None
    brand = _extract_value(text, ["бренд", "brand"]) or "Нет бренда"
    category = _extract_value(text, ["категория", "category"]) or _guess_category(text)
    if category:
        category = _map_to_ozon_category_path(category)
    dimensions = _extract_dimensions(text)
    weight = _extract_weight(text)
    title = _build_title(text, brand, category)
    description = _build_description(text, title)
    attributes = _extract_attributes(text)
    keywords = _build_keywords(title, category, attributes, brand)
    selling_points = _build_selling_points(title)

    return OzonCardDraft(
        offer_id=_build_offer_id(title),
        name=title,
        description=description,
        category_hint=category,
        brand=brand,
        price=price,
        old_price=old_price,
        weight_g=weight,
        width_mm=dimensions.get("width_mm"),
        height_mm=dimensions.get("height_mm"),
        depth_mm=dimensions.get("depth_mm"),
        images=[url for url in urls if _looks_like_image(url)],
        attributes=attributes,
        keywords=keywords,
        selling_points=selling_points,
        checklist=_build_checklist(price, urls, weight, dimensions),
        source_text=task.strip(),
    )


async def build_enhanced_ozon_card_draft(
    task: str,
    competitors: list[dict] | None = None,
    profile: dict | None = None,
) -> OzonCardDraft:
    draft = build_ozon_card_draft(task)
    competitors = competitors or []
    profile = profile or {}
    apply_card_profile(draft, profile)
    _apply_competitor_context(draft, competitors)
    if not ai_is_available():
        return draft

    response = await ask_ai(
        _build_ai_card_prompt(draft, competitors, profile),
        system=AI_CARD_SYSTEM_PROMPT,
        max_tokens=1600,
    )
    if _is_ai_error(response):
        draft.ai_notes = "AI-доработка недоступна, черновик собран локально с учетом конкурентов."
        return draft
    data = _extract_json_object(response)
    if not data:
        draft.ai_notes = _shorten(response, 1200)
        return draft

    draft.name = _clean_ai_string(data.get("name"), max_len=140) or draft.name
    draft.description = _clean_ai_string(data.get("description"), max_len=4000) or draft.description
    draft.category_hint = _clean_ai_string(data.get("category_hint"), max_len=200) or draft.category_hint
    draft.brand = _clean_ai_string(data.get("brand"), max_len=120) or draft.brand
    draft.attributes.update(_clean_string_dict(data.get("attributes")))
    draft.keywords = _sanitize_keywords(
        _clean_string_list(data.get("keywords"), limit=30) or draft.keywords,
        brand=draft.brand,
    )
    draft.selling_points = _clean_string_list(data.get("selling_points"), limit=4) or draft.selling_points
    draft.checklist = _clean_string_list(data.get("checklist"), limit=12) or draft.checklist
    draft.ai_notes = _clean_ai_string(data.get("notes"), max_len=1200) or "AI улучшил черновик карточки."
    apply_card_profile(draft, profile)
    return draft


def build_ozon_import_payload(draft: OzonCardDraft) -> dict:
    item = {
        "offer_id": draft.offer_id,
        "name": draft.name,
        "description": draft.description,
        "barcode": "",
        "price": str(draft.price or ""),
        "old_price": str(draft.old_price or ""),
        "vat": draft.vat,
        "currency_code": "RUB",
        "images": draft.images,
        "weight": draft.weight_g,
        "weight_unit": "g",
        "width": draft.width_mm,
        "height": draft.height_mm,
        "depth": draft.depth_mm,
        "dimension_unit": "mm",
        "attributes": [
            {"name": key, "value": value}
            for key, value in sorted(draft.attributes.items())
        ],
    }
    return {
        "endpoint_note": "Черновик под Ozon Seller API. Перед отправкой нужны category_id/type_id и обязательные attribute_id из Ozon.",
        "method": "POST /v3/product/import",
        "items": [item],
    }


def build_ozon_card_search_query(draft: OzonCardDraft) -> str:
    query = _strip_commercial_terms(draft.name)
    if not query and draft.category_hint:
        query = _strip_commercial_terms(draft.category_hint)
    return query or draft.name


def export_ozon_card_json(draft: OzonCardDraft) -> io.BytesIO:
    payload = {
        "draft": asdict(draft),
        "ozon_import_payload": build_ozon_import_payload(draft),
    }
    buf = io.BytesIO(json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.name = f"ozon_card_{draft.offer_id}.json"
    buf.seek(0)
    return buf


def export_ozon_card_xlsx(draft: OzonCardDraft) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Карточка Ozon"

    ws["A1"] = "Черновик карточки Ozon"
    ws["A1"].font = Font(size=16, bold=True, color="1F2937")
    ws["A2"] = f"Сгенерировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(color="6B7280")
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")

    rows = [
        ("Артикул", draft.offer_id, "Уникальный offer_id"),
        ("Название", draft.name, "Тип + бренд + ключевые свойства"),
        ("Категория", draft.category_hint or "", "Проверьте точную категорию Ozon"),
        ("Бренд", draft.brand, "Для неизвестного бренда обычно указывают 'Нет бренда'"),
        ("Цена", draft.price or "", "Цена продажи, RUB"),
        ("Старая цена", draft.old_price or "", "Можно очистить, если не нужна"),
        ("НДС", draft.vat, "0, 10, 20 или без НДС согласно кабинету"),
        ("Вес, г", draft.weight_g or "", "Обязательное поле для API"),
        ("Ширина, мм", draft.width_mm or "", "Обязательное поле для API"),
        ("Высота, мм", draft.height_mm or "", "Обязательное поле для API"),
        ("Глубина, мм", draft.depth_mm or "", "Обязательное поле для API"),
        ("Фото", "\n".join(draft.images), "URL изображений"),
        ("Ключевые слова", ", ".join(draft.keywords), "Для SEO и проверки полноты"),
        ("Чек-лист", "\n".join(draft.checklist), "Что проверить перед публикацией"),
        ("Конкуренты", draft.competitor_summary, "Краткий анализ выдачи Ozon"),
        ("AI-заметки", draft.ai_notes, "Пояснения и ограничения черновика"),
        ("Описание", draft.description, "Можно редактировать перед загрузкой"),
    ]

    try:
        ws.unmerge_cells("A1:C1")
        ws.unmerge_cells("A2:C2")
    except Exception:
        pass
    ws["A1"] = ""
    ws["A2"] = ""
    rows.insert(14, ("Преимущества", "\n".join(draft.selling_points), "Короткие буллиты с выгодами для покупателя"))

    start = 1
    headers = ["Поле", "Значение", "Комментарий"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(rows, start + 1):
        for col, value in enumerate(row, 1):
            cell = ws.cell(idx, col, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    attr_start = start + len(rows) + 3
    ws.cell(attr_start, 1, "Характеристики").font = Font(size=14, bold=True)
    ws.cell(attr_start + 1, 1, "Название").font = Font(bold=True, color="FFFFFF")
    ws.cell(attr_start + 1, 2, "Значение").font = Font(bold=True, color="FFFFFF")
    ws.cell(attr_start + 1, 1).fill = PatternFill("solid", fgColor="059669")
    ws.cell(attr_start + 1, 2).fill = PatternFill("solid", fgColor="059669")

    for row_idx, (key, value) in enumerate(sorted(draft.attributes.items()), attr_start + 2):
        ws.cell(row_idx, 1, key)
        ws.cell(row_idx, 2, value)

    widths = {1: 22, 2: 64, 3: 44}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.name = f"ozon_card_{draft.offer_id}.xlsx"
    buf.seek(0)
    return buf


def export_ozon_cards_batch_json(items: list[OzonCardBatchItem]) -> io.BytesIO:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "summary": _build_batch_counts(items),
        "items": [
            {
                "source": item.source,
                "status": item.status,
                "message": item.message,
                "draft": asdict(item.draft) if item.draft else None,
                "ozon_import_payload": build_ozon_import_payload(item.draft) if item.draft else None,
            }
            for item in items
        ],
    }
    buf = io.BytesIO(json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"))
    buf.name = f"ozon_cards_batch_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
    buf.seek(0)
    return buf


def export_ozon_cards_batch_xlsx(items: list[OzonCardBatchItem]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    ws_cards = wb.create_sheet("Карточки")
    ws_attrs = wb.create_sheet("Характеристики")

    counts = _build_batch_counts(items)
    ws_summary["A1"] = "Пакет карточек Ozon"
    ws_summary["A1"].font = Font(size=16, bold=True, color="1F2937")
    ws_summary["A2"] = f"Сгенерировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_summary["A3"] = f"Всего: {counts['total']} | готово к проверке: {counts['ready']} | нужна доработка: {counts['needs_review']} | ошибки: {counts['error']}"
    ws_summary["A5"] = "Как работать с файлом"
    ws_summary["A5"].font = Font(bold=True)
    ws_summary["A6"] = "1. Начните с листа «Карточки» и отфильтруйте статус «Нужна доработка»."
    ws_summary["A7"] = "2. Заполните пропуски: категорию, фото, вес и габариты."
    ws_summary["A8"] = "3. Лист «Характеристики» используйте как подсказку для обязательных полей Ozon."

    for col, width in {1: 28, 2: 28, 3: 80}.items():
        ws_summary.column_dimensions[get_column_letter(col)].width = width

    card_headers = [
        "№", "Статус", "Что проверить", "Артикул", "Название", "Категория", "Бренд",
        "Цена", "Старая цена", "НДС", "Вес, г", "Ширина, мм", "Высота, мм", "Глубина, мм",
        "Фото", "Ключевые слова", "Описание", "Конкуренты", "AI-заметки", "Источник", "Ошибка",
    ]
    _write_header_row(ws_cards, card_headers)
    for idx, item in enumerate(items, 1):
        draft = item.draft
        missing = _draft_missing_fields(draft) if draft else []
        row = [
            idx,
            _human_batch_status(item.status),
            ", ".join(missing) if missing else "",
            draft.offer_id if draft else "",
            draft.name if draft else "",
            draft.category_hint if draft else "",
            draft.brand if draft else "",
            draft.price if draft else "",
            draft.old_price if draft else "",
            draft.vat if draft else "",
            draft.weight_g if draft else "",
            draft.width_mm if draft else "",
            draft.height_mm if draft else "",
            draft.depth_mm if draft else "",
            "\n".join(draft.images) if draft else "",
            ", ".join(draft.keywords) if draft else "",
            draft.description if draft else "",
            draft.competitor_summary if draft else "",
            draft.ai_notes if draft else "",
            item.source,
            item.message,
        ]
        ws_cards.append(row)

    attr_headers = ["№", "Артикул", "Название", "Характеристика", "Значение"]
    _write_header_row(ws_attrs, attr_headers)
    for idx, item in enumerate(items, 1):
        if not item.draft:
            continue
        for key, value in sorted(item.draft.attributes.items()):
            ws_attrs.append([idx, item.draft.offer_id, item.draft.name, key, value])

    _style_batch_sheet(ws_cards)
    _style_batch_sheet(ws_attrs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.name = f"ozon_cards_batch_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    buf.seek(0)
    return buf


def format_ozon_card_preview(draft: OzonCardDraft) -> str:
    missing = []
    if not draft.category_hint:
        missing.append("категория")
    if not draft.price:
        missing.append("цена")
    if not draft.images:
        missing.append("фото")
    if not draft.weight_g:
        missing.append("вес")
    if not (draft.width_mm and draft.height_mm and draft.depth_mm):
        missing.append("габариты")

    missing_text = ", ".join(missing) if missing else "критичных пропусков нет"
    readiness = "можно проверять и дорабатывать" if not missing else "нужна ручная доработка перед загрузкой"
    next_steps = (
        "Что делать дальше:\n"
        "1. Откройте XLSX и заполните/проверьте пункты из строки «Что проверить».\n"
        "2. Сверьте категорию и обязательные атрибуты в кабинете Ozon Seller.\n"
        "3. После правок используйте XLSX как черновик для ручного импорта. JSON пока нужен как техфайл для будущей отправки через Seller API."
    )
    return (
        "<b>Черновик карточки Ozon готов</b>\n\n"
        f"<b>Название:</b> {draft.name}\n"
        f"<b>Категория:</b> {draft.category_hint or 'нужно уточнить'}\n"
        f"<b>Артикул:</b> <code>{draft.offer_id}</code>\n"
        f"<b>Цена:</b> {draft.price or 'не указана'} ₽\n"
        f"<b>Бренд:</b> {draft.brand}\n"
        f"<b>Характеристик:</b> {len(draft.attributes)}\n"
        f"<b>Фото:</b> {len(draft.images)}\n"
        f"<b>Статус:</b> {readiness}\n"
        f"<b>Что проверить:</b> {missing_text}\n\n"
        f"{next_steps}"
    )


def format_ozon_batch_preview(items: list[OzonCardBatchItem]) -> str:
    counts = _build_batch_counts(items)
    problem_lines = []
    for idx, item in enumerate(items, 1):
        if item.status == "error":
            problem_lines.append(f"{idx}. ошибка: {item.message[:80]}")
            continue
        missing = _draft_missing_fields(item.draft)
        if missing:
            name = item.draft.name[:60] if item.draft else item.source[:60]
            problem_lines.append(f"{idx}. {name}: {', '.join(missing)}")

    problems_text = "\n".join(problem_lines[:10]) if problem_lines else "критичных пропусков нет"
    if len(problem_lines) > 10:
        problems_text += f"\n...и ещё {len(problem_lines) - 10}"

    return (
        "<b>Пакет карточек Ozon готов</b>\n\n"
        f"<b>Всего:</b> {counts['total']}\n"
        f"<b>Можно проверять:</b> {counts['ready']}\n"
        f"<b>Нужна доработка:</b> {counts['needs_review']}\n"
        f"<b>Ошибки:</b> {counts['error']}\n\n"
        f"<b>Что требует внимания:</b>\n{problems_text}\n\n"
        "Что дальше: откройте общий XLSX, начните с листа «Сводка», затем на листе «Карточки» "
        "отфильтруйте «Нужна доработка». JSON прикладываю как технический файл под будущую API-загрузку."
    )


def _build_title(text: str, brand: str, category: str | None) -> str:
    explicit = _extract_value(text, ["название", "товар", "name"])
    if explicit:
        return _strip_commercial_terms(explicit)[:140]

    first_line = next((line.strip() for line in text.splitlines() if line.strip()), text)
    first_line = re.sub(r"https?://\S+", "", first_line)
    first_line = re.sub(r"\b(цена|стоимость|бренд|категория)\s*[:=-].*", "", first_line, flags=re.I)
    first_line = _strip_commercial_terms(first_line)
    words = [w for w in re.split(r"\s+", first_line) if len(w) > 1]
    title = " ".join(words[:16]).strip(" ,.;:-")
    if not title and category:
        title = category
    if brand != "Нет бренда" and brand.lower() not in title.lower():
        title = f"{title} {brand}".strip()
    return title[:140] or "Новый товар"


def _build_description(text: str, title: str) -> str:
    explicit = _extract_value(text, ["описание", "description"])
    if explicit:
        base = explicit
    else:
        clean = re.sub(r"https?://\S+", "", text)
        base = _clean(clean)

    if len(base) < 80:
        base = (
            f"{title}. Подходит для ежедневного использования. "
            "Перед публикацией проверьте комплектацию, совместимость, материалы и гарантийные условия."
        )
    return base[:4000]


def _extract_attributes(text: str) -> dict[str, str]:
    attributes: dict[str, str] = {}
    for line in text.splitlines():
        match = re.match(r"\s*([A-Za-zА-Яа-яЁё0-9 /_-]{3,40})\s*[:=-]\s*(.{2,120})\s*$", line)
        if not match:
            continue
        key = match.group(1).strip().capitalize()
        value = match.group(2).strip()
        if key.lower() in {"цена", "стоимость", "название", "товар", "описание", "бренд", "категория"}:
            continue
        attributes[key] = value

    color = _extract_value(text, ["цвет"])
    if color:
        attributes.setdefault("Цвет", color)
    material = _extract_value(text, ["материал"])
    if material:
        attributes.setdefault("Материал", material)
    package = _extract_value(text, ["комплектация"])
    if package:
        attributes.setdefault("Комплектация", package)
    return attributes


def _build_keywords(title: str, category: str | None, attributes: dict[str, str], brand: str) -> list[str]:
    source = " ".join([title, category or "", " ".join(attributes.values())])
    words = re.findall(r"[A-Za-zА-Яа-яЁё0-9+_-]{3,}", source.lower())
    cleaned = _sanitize_keywords(words, brand=brand, limit=30)
    whitelist_roots = _category_keyword_whitelist(category)
    if whitelist_roots:
        scoped = [w for w in cleaned if any(root in w for root in whitelist_roots)]
        if scoped:
            return scoped[:20]
    return cleaned[:20]


def _extract_value(text: str, keys: list[str]) -> str | None:
    for key in keys:
        match = re.search(rf"(?:^|\n)\s*{re.escape(key)}\s*[:=-]\s*(.+)", text, flags=re.I)
        if match:
            return match.group(1).strip()[:300]
    return None


def _extract_price(text: str) -> int | None:
    patterns = [
        r"(?:цена|стоимость)\s*[:=-]?\s*([\d\s]{2,8})(?:₽|руб|р)?",
        r"(?:за|по)\s*([\d\s]{2,8})(?:₽|руб|р)?",
        r"([\d\s]{2,8})\s*(?:₽|руб|р)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            value = int(re.sub(r"\D", "", match.group(1)) or "0")
            if 1 <= value <= 50_000_000:
                return value
    return None


def _strip_commercial_terms(text: str) -> str:
    text = re.sub(r"\b(?:цена|стоимость)\s*[:=-]?\s*[\d\s]{1,8}\s*(?:₽|руб\.?|р\.?)?", " ", text, flags=re.I)
    text = re.sub(r"\b(?:за|по)\s*[\d\s]{2,8}\s*(?:₽|руб\.?|р\.?)?", " ", text, flags=re.I)
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,.;:-")


def _extract_dimensions(text: str) -> dict[str, int]:
    match = re.search(r"(\d{1,4})\s*[xх*]\s*(\d{1,4})\s*[xх*]\s*(\d{1,4})\s*(мм|см|mм)?", text, flags=re.I)
    if not match:
        return {}
    values = [int(match.group(i)) for i in range(1, 4)]
    unit = (match.group(4) or "мм").lower()
    if unit == "см":
        values = [v * 10 for v in values]
    return {"width_mm": values[0], "height_mm": values[1], "depth_mm": values[2]}


def _extract_weight(text: str) -> int | None:
    match = re.search(r"(?:вес|weight)\s*[:=-]?\s*(\d{1,6})\s*(кг|kg|г|g)?", text, flags=re.I)
    if not match:
        return None
    value = int(match.group(1))
    unit = (match.group(2) or "г").lower()
    if unit in {"кг", "kg"}:
        value *= 1000
    return value


def _extract_urls(text: str) -> list[str]:
    return [url.rstrip(".,);]}>") for url in re.findall(r"https?://\S+", text or "", flags=re.I)]


def _looks_like_image(url: str) -> bool:
    parsed = urlparse(url)
    path = parsed.path.lower()
    return path.endswith((".jpg", ".jpeg", ".png", ".webp"))


def _guess_category(text: str) -> str | None:
    low = text.lower()
    rules = [
        ("держател", "Автомобильные держатели"),
        ("наушник", "Наушники и гарнитуры"),
        ("кабель", "Кабели"),
        ("чехол", "Чехлы"),
        ("заряд", "Зарядные устройства"),
        ("подписка", "Цифровые товары"),
    ]
    for needle, category in rules:
        if needle in low:
            return _map_to_ozon_category_path(category)
    wb_category = _extract_value(text, ["категория wb", "wb category", "категория"])
    if wb_category:
        return _map_to_ozon_category_path(wb_category)
    return None


def _build_offer_id(title: str) -> str:
    translit = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh", "з": "z",
        "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
        "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh",
        "щ": "sch", "ы": "y", "э": "e", "ю": "yu", "я": "ya",
    }
    raw = "".join(translit.get(ch, ch) for ch in title.lower())
    slug = re.sub(r"[^a-z0-9]+", "-", raw).strip("-")[:32] or "product"
    digest = hashlib.md5(title.encode("utf-8")).hexdigest()[:6]
    return f"{slug}-{digest}"[:50]


def _clean(text: str) -> str:
    return re.sub(r"[ \t]+", " ", text or "").strip()


AI_CARD_SYSTEM_PROMPT = """Ты эксперт по карточкам товаров для Ozon и Wildberries.
Улучши черновик карточки, но не выдумывай факты, которых нет в исходных данных.
Пиши на русском, без медицинских, юридических, гарантийных обещаний и без абсолютов вроде "лучший".
Верни только валидный JSON без markdown-блока."""


def _build_ai_card_prompt(draft: OzonCardDraft, competitors: list[dict], profile: dict | None = None) -> str:
    profile = profile or {}
    payload = {
        "source_text": draft.source_text,
        "draft": asdict(draft),
        "profile": {
            "language": profile.get("language", "ru"),
            "tone_voice": profile.get("tone_voice", "professional"),
            "max_length": profile.get("max_length", 1500),
            "forbidden_words": profile.get("forbidden_words", []),
        },
        "competitors": _compact_competitors(competitors),
        "competitor_summary": draft.competitor_summary,
        "required_json_schema": {
            "name": "SEO-название до 140 символов",
            "description": "продающее описание 700-1500 символов",
            "category_hint": "предполагаемая категория",
            "brand": "бренд или Нет бренда",
            "attributes": {"Название характеристики": "Значение"},
            "keywords": ["ключевые слова без повторов"],
            "selling_points": ["3-4 коротких буллита с выгодами для покупателя"],
            "checklist": ["что проверить или добавить перед публикацией"],
            "notes": "коротко: что улучшено и какие есть ограничения",
        },
    }
    return (
        "Улучши карточку товара для маркетплейса. "
        "Сохрани фактические данные из исходника. Конкурентов используй только для SEO-слов, ценового ориентира и чек-листа, "
        "не копируй чужие названия дословно и не выдумывай свойства товара. "
        "Если данных не хватает, вынеси это в checklist.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )


def _extract_json_object(text: str) -> dict | None:
    if not text:
        return None
    raw = text.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.I)
        raw = re.sub(r"\s*```$", "", raw)
    start = raw.find("{")
    end = raw.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return None
    try:
        data = json.loads(raw[start:end + 1])
    except Exception as e:
        logger.warning(f"AI card JSON parse failed: {e}")
        return None
    return data if isinstance(data, dict) else None


def _clean_ai_string(value, *, max_len: int) -> str | None:
    if not isinstance(value, str):
        return None
    value = _clean(value)
    return value[:max_len] if value else None


def _clean_string_list(value, *, limit: int) -> list[str]:
    if not isinstance(value, list):
        return []
    result = []
    seen = set()
    for item in value:
        if not isinstance(item, str):
            continue
        item = _clean(item)
        marker = item.lower()
        if item and marker not in seen:
            result.append(item[:160])
            seen.add(marker)
        if len(result) >= limit:
            break
    return result


def _clean_string_dict(value) -> dict[str, str]:
    if not isinstance(value, dict):
        return {}
    result = {}
    for key, item in value.items():
        if not isinstance(key, str) or not isinstance(item, str):
            continue
        key = _clean(key).capitalize()[:80]
        item = _clean(item)[:300]
        if key and item:
            result[key] = item
    return result


def _shorten(text: str, limit: int) -> str:
    return _clean(text)[:limit]


def _is_ai_error(text: str) -> bool:
    if not text:
        return True
    low = text.lower()
    return low.startswith("ai error:") or "ai недоступен" in low or "ai unavailable" in low


def apply_card_profile(draft: OzonCardDraft, profile: dict | None) -> OzonCardDraft:
    profile = profile or {}
    forbidden = [str(x).strip().lower() for x in profile.get("forbidden_words", []) if str(x).strip()]
    if forbidden:
        draft.name = _remove_forbidden_words(draft.name, forbidden)
        draft.description = _remove_forbidden_words(draft.description, forbidden)
        draft.keywords = [k for k in draft.keywords if k.lower() not in forbidden]
        draft.selling_points = [_remove_forbidden_words(point, forbidden) for point in draft.selling_points]
        draft.selling_points = [x for x in draft.selling_points if x]

    max_len = int(profile.get("max_length") or 1500)
    if max_len > 0:
        draft.description = draft.description[:max_len]

    for attr in profile.get("required_attributes", []) or []:
        name = str(attr).strip()
        if not name or name in {"Бренд", "Категория"}:
            continue
        draft.attributes.setdefault(name, "нужно заполнить")

    return draft


def _remove_forbidden_words(text: str, forbidden: list[str]) -> str:
    value = text or ""
    for word in forbidden:
        if not word:
            continue
        value = re.sub(rf"\b{re.escape(word)}\b", " ", value, flags=re.I)
    value = re.sub(r"\s{2,}", " ", value).strip(" ,.;:-")
    return value


def _build_checklist(price: int | None, urls: list[str], weight: int | None, dimensions: dict[str, int]) -> list[str]:
    checklist = []
    if not price:
        checklist.append("Указать цену продажи")
    if not any(_looks_like_image(url) for url in urls):
        checklist.append("Добавить фото товара на белом фоне и дополнительные ракурсы")
    if not weight:
        checklist.append("Проверить вес в граммах")
    if not (dimensions.get("width_mm") and dimensions.get("height_mm") and dimensions.get("depth_mm")):
        checklist.append("Проверить габариты упаковки в миллиметрах")
    checklist.append("Сверить точную категорию и обязательные атрибуты в кабинете Ozon")
    return checklist


def _draft_missing_fields(draft: OzonCardDraft | None) -> list[str]:
    if not draft:
        return ["карточка не собрана"]
    missing = []
    if not draft.category_hint:
        missing.append("категория")
    if not draft.price:
        missing.append("цена")
    if not draft.images:
        missing.append("фото")
    if not draft.weight_g:
        missing.append("вес")
    if not (draft.width_mm and draft.height_mm and draft.depth_mm):
        missing.append("габариты")
    return missing


def _build_batch_counts(items: list[OzonCardBatchItem]) -> dict[str, int]:
    counts = {"total": len(items), "ready": 0, "needs_review": 0, "error": 0}
    for item in items:
        if item.status == "error":
            counts["error"] += 1
        elif _draft_missing_fields(item.draft):
            counts["needs_review"] += 1
        else:
            counts["ready"] += 1
    return counts


def _human_batch_status(status: str) -> str:
    if status == "ready":
        return "Готово к проверке"
    if status == "needs_review":
        return "Нужна доработка"
    if status == "error":
        return "Ошибка"
    return status or "Неизвестно"


def _write_header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_batch_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    widths = {
        1: 6,
        2: 20,
        3: 28,
        4: 28,
        5: 52,
        6: 28,
        7: 18,
        8: 12,
        15: 48,
        16: 44,
        17: 70,
        18: 52,
        19: 52,
        20: 60,
        21: 42,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _apply_competitor_context(draft: OzonCardDraft, competitors: list[dict]) -> None:
    if not competitors:
        return

    prices = [item["price"] for item in competitors if isinstance(item.get("price"), int)]
    words = _top_competitor_words(competitors)
    for word, _ in words[:10]:
        if word not in draft.keywords:
            draft.keywords.append(word)
    draft.keywords = _sanitize_keywords(draft.keywords, brand=draft.brand, limit=30)

    parts = [f"Найдено конкурентов: {len(competitors)}"]
    if prices:
        parts.append(f"цены: {min(prices)}-{max(prices)} ₽, медиана {int(median(prices))} ₽")
        if not draft.price:
            draft.checklist.append(f"Сравнить цену с медианой конкурентов: {int(median(prices))} ₽")
    if words:
        parts.append("частые слова: " + ", ".join(word for word, _ in words[:8]))
    draft.competitor_summary = "; ".join(parts)
    if "Проверить, чем карточка отличается от конкурентов в выдаче Ozon" not in draft.checklist:
        draft.checklist.append("Проверить, чем карточка отличается от конкурентов в выдаче Ozon")


def _compact_competitors(competitors: list[dict]) -> list[dict]:
    return [
        {
            "name": (item.get("name") or "")[:120],
            "price": item.get("price"),
            "url": item.get("url"),
        }
        for item in competitors[:10]
    ]


def _top_competitor_words(competitors: list[dict]) -> list[tuple[str, int]]:
    stop = {
        "для", "или", "без", "при", "под", "над", "товар", "шт", "набор",
        "ozon", "wb", "wildberries", "черный", "белый",
    }
    counts: dict[str, int] = {}
    for item in competitors:
        name = item.get("name") or ""
        for word in re.findall(r"[A-Za-zА-Яа-яЁё0-9+]{3,}", name.lower()):
            if word in stop or word.isdigit():
                continue
            counts[word] = counts.get(word, 0) + 1
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:20]


TECH_SEO_TERMS = {
    "https", "http", "www", "webp", "images", "image", "img", "jpg", "jpeg", "png",
    "catalog", "wildberries", "ozon", "basket", "vol", "ru", "product", "api",
}
RU_STOP_WORDS = {"для", "или", "без", "это", "под", "при", "товар", "цена", "руб", "за"}


def _sanitize_keywords(words: list[str], brand: str, limit: int = 30) -> list[str]:
    brand_tokens = {w.lower() for w in re.findall(r"[A-Za-zА-Яа-яЁё]{2,}", brand or "")}
    result: list[str] = []
    seen: set[str] = set()
    for raw in words:
        word = _clean(raw).lower().strip("-_")
        if not word or word in seen:
            continue
        if any(ch.isdigit() for ch in word):
            continue
        if word in TECH_SEO_TERMS or word in RU_STOP_WORDS:
            continue
        has_latin = bool(re.search(r"[a-z]", word))
        has_cyr = bool(re.search(r"[а-яё]", word))
        if has_latin and word not in brand_tokens:
            continue
        if not has_cyr and word not in brand_tokens:
            continue
        result.append(word)
        seen.add(word)
        if len(result) >= limit:
            break
    return result


def _build_selling_points(title: str) -> list[str]:
    base = title.strip() or "Товар"
    return [
        f"{base}: удобен для ежедневного использования",
        "Практичный вариант для дома и работы",
        "Продуманный формат и понятное применение",
        "Хорошо подходит как покупка для себя и в подарок",
    ]


def _map_to_ozon_category_path(category: str) -> str:
    low = (category or "").lower()
    mapping = [
        ("мебел", "Дом и сад/Мебель/Кресла"),
        ("кресл", "Дом и сад/Мебель/Кресла"),
        ("коврик", "Спорт и отдых/Фитнес и йога/Коврики для йоги"),
        ("йог", "Спорт и отдых/Фитнес и йога/Коврики для йоги"),
        ("кабел", "Электроника/Аксессуары для смартфонов/Кабели и переходники"),
        ("держател", "Автотовары/Аксессуары/Держатели для телефонов"),
        ("наушник", "Электроника/Аудиотехника/Наушники и гарнитуры"),
        ("чехол", "Электроника/Аксессуары для смартфонов/Чехлы"),
    ]
    for token, ozon_path in mapping:
        if token in low:
            return ozon_path
    return category


def _category_keyword_whitelist(category: str | None) -> list[str]:
    low = (category or "").lower()
    rules: list[tuple[tuple[str, ...], list[str]]] = [
        (("коврик", "йога", "фитнес"), ["коврик", "йог", "фитнес", "нескольз", "толщин", "упражнен"]),
        (("кабел", "смартфон", "аксессуар"), ["кабел", "заряд", "usb", "type", "lightning", "micro", "разъем"]),
        (("мебел", "кресл"), ["кресл", "эргоном", "офис", "обивк", "подлокот", "спинк"]),
        (("держател", "автотовар"), ["держател", "телефон", "магнит", "панел", "авто"]),
        (("наушник", "аудио"), ["науш", "гарнитур", "звук", "шумоподав", "микрофон"]),
        (("чехл",), ["чехл", "защит", "смартфон", "материал", "удар"]),
        (("маникюр", "педикюр", "красота"), ["кусач", "маникюр", "педикюр", "лезви", "ногт"]),
    ]
    for triggers, roots in rules:
        if any(trigger in low for trigger in triggers):
            return roots
    return []

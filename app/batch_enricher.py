import csv
import json
from dataclasses import asdict
from pathlib import Path
from typing import Any

import openpyxl

from app.card_filler import build_ozon_card_draft, build_ozon_import_payload


CHECKPOINT_SUFFIX = ".checkpoint.json"


def enrich_file(
    input_path: str | Path,
    output_path: str | Path,
    *,
    limit: int | None = None,
    resume: bool = False,
) -> dict[str, Any]:
    input_path = Path(input_path)
    output_path = Path(output_path)
    checkpoint_path = output_path.with_suffix(output_path.suffix + CHECKPOINT_SUFFIX)

    rows = _read_rows(input_path)
    processed_ids = _load_checkpoint(checkpoint_path) if resume else set()

    output_rows: list[dict[str, Any]] = []
    processed_now = 0
    skipped = 0

    for index, row in enumerate(rows, 1):
        row_id = _row_id(index, row)
        if row_id in processed_ids:
            skipped += 1
            output_rows.append(row)
            continue
        if limit is not None and processed_now >= limit:
            output_rows.append(row)
            continue

        enriched = enrich_row(row)
        output_rows.append(enriched)
        processed_ids.add(row_id)
        processed_now += 1
        _save_checkpoint(checkpoint_path, processed_ids)

    _write_xlsx(output_path, output_rows)
    return {
        "input": str(input_path),
        "output": str(output_path),
        "checkpoint": str(checkpoint_path),
        "total_rows": len(rows),
        "processed": processed_now,
        "skipped": skipped,
        "resume": resume,
    }


def enrich_row(row: dict[str, Any]) -> dict[str, Any]:
    task = _row_to_task(row)
    draft = build_ozon_card_draft(task)
    payload = build_ozon_import_payload(draft)
    enriched = dict(row)
    enriched.update({
        "agent_offer_id": draft.offer_id,
        "agent_name": draft.name,
        "agent_description": draft.description,
        "agent_category_hint": draft.category_hint or "",
        "agent_brand": draft.brand,
        "agent_price": draft.price or "",
        "agent_old_price": draft.old_price or "",
        "agent_weight_g": draft.weight_g or "",
        "agent_width_mm": draft.width_mm or "",
        "agent_height_mm": draft.height_mm or "",
        "agent_depth_mm": draft.depth_mm or "",
        "agent_images": "\n".join(draft.images),
        "agent_keywords": ", ".join(draft.keywords),
        "agent_attributes_json": json.dumps(draft.attributes, ensure_ascii=False, sort_keys=True),
        "agent_ozon_payload_json": json.dumps(payload["items"][0], ensure_ascii=False, sort_keys=True),
        "agent_needs_review": _needs_review(draft),
    })
    return enriched


def _read_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError(f"Unsupported input format: {path.suffix}. Use .xlsx or .csv")


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else f"column_{idx}" for idx, value in enumerate(rows[0], 1)]
    result = []
    for values in rows[1:]:
        if not any(value not in (None, "") for value in values):
            continue
        result.append({headers[idx]: values[idx] if idx < len(values) else "" for idx in range(len(headers))})
    return result


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True) if path.parent != Path("") else None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "enriched"
    headers = _collect_headers(rows)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for cell in ws[1]:
        cell.style = "Headline 3"
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 2, 14), 55)
        ws.column_dimensions[column_cells[0].column_letter].width = width
    wb.save(path)


def _collect_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    return headers


def _row_to_task(row: dict[str, Any]) -> str:
    priority = [
        "товар", "название", "name", "title", "product", "описание", "description",
        "бренд", "brand", "категория", "category", "цена", "price",
    ]
    lines = []
    used = set()
    lower_map = {str(key).lower(): key for key in row}
    for key in priority:
        original = lower_map.get(key)
        if original is not None and row.get(original) not in (None, ""):
            lines.append(f"{key}: {row[original]}")
            used.add(original)
    for key, value in row.items():
        if key in used or value in (None, ""):
            continue
        lines.append(f"{key}: {value}")
    return "\n".join(lines)


def _needs_review(draft) -> str:
    missing = []
    if not draft.price:
        missing.append("price")
    if not draft.category_hint:
        missing.append("category")
    if not draft.images:
        missing.append("images")
    if not draft.weight_g:
        missing.append("weight")
    if not (draft.width_mm and draft.height_mm and draft.depth_mm):
        missing.append("dimensions")
    return ", ".join(missing)


def _row_id(index: int, row: dict[str, Any]) -> str:
    for key in ("id", "ID", "sku", "SKU", "артикул", "Артикул"):
        value = row.get(key)
        if value not in (None, ""):
            return f"{key}:{value}"
    return f"row:{index}"


def _load_checkpoint(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return set(payload.get("processed_ids", []))
    except Exception:
        return set()


def _save_checkpoint(path: Path, processed_ids: set[str]) -> None:
    path.write_text(
        json.dumps({"processed_ids": sorted(processed_ids)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

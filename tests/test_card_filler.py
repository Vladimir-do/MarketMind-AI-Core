import json
import unittest
from zipfile import ZipFile
from unittest.mock import AsyncMock, patch
import openpyxl

from app.card_filler import (
    OzonCardBatchItem,
    apply_card_profile,
    build_enhanced_ozon_card_draft,
    build_ozon_card_draft,
    build_ozon_card_search_query,
    build_ozon_import_payload,
    export_ozon_cards_batch_json,
    export_ozon_cards_batch_xlsx,
    export_ozon_card_json,
    export_ozon_card_xlsx,
    format_ozon_batch_preview,
    format_ozon_card_preview,
)


class CardFillerTests(unittest.IsolatedAsyncioTestCase):
    def test_build_ozon_card_draft_from_task(self):
        draft = build_ozon_card_draft(
            """
            товар: Держатель телефона автомобильный
            бренд: Нет бренда
            категория: Автомобильные держатели
            цена: 599
            цвет: черный
            материал: ABS пластик
            вес: 180 г
            размер: 120x80x70 мм
            комплектация: держатель, крепление, коробка
            https://example.com/photo.jpg
            """
        )

        self.assertIn("Держатель телефона", draft.name)
        self.assertEqual(draft.category_hint, "Автотовары/Аксессуары/Держатели для телефонов")
        self.assertEqual(draft.price, 599)
        self.assertEqual(draft.weight_g, 180)
        self.assertEqual(draft.width_mm, 120)
        self.assertEqual(draft.height_mm, 80)
        self.assertEqual(draft.depth_mm, 70)
        self.assertEqual(draft.attributes["Цвет"], "черный")
        self.assertEqual(draft.images, ["https://example.com/photo.jpg"])

    def test_build_ozon_import_payload_contains_item(self):
        draft = build_ozon_card_draft("товар: Кабель USB-C\nцена: 299\nвес: 50 г\nразмер: 100x80x20 мм")
        payload = build_ozon_import_payload(draft)

        self.assertEqual(payload["method"], "POST /v3/product/import")
        self.assertEqual(payload["items"][0]["offer_id"], draft.offer_id)
        self.assertEqual(payload["items"][0]["currency_code"], "RUB")

    def test_card_draft_separates_inline_price_from_title(self):
        draft = build_ozon_card_draft("кусачки маникюрные цена 499")

        self.assertEqual(draft.price, 499)
        self.assertNotIn("цена 499", draft.name.lower())
        self.assertEqual(build_ozon_card_search_query(draft), "кусачки маникюрные")

    def test_card_draft_extracts_price_after_za(self):
        draft = build_ozon_card_draft("коврик для йоги за 1200р")

        self.assertEqual(draft.price, 1200)
        self.assertNotIn("1200", draft.name)

    def test_exports_json_and_xlsx(self):
        draft = build_ozon_card_draft("товар: Кабель USB-C\nцена: 299\nвес: 50 г\nразмер: 100x80x20 мм")
        json_buf = export_ozon_card_json(draft)
        xlsx_buf = export_ozon_card_xlsx(draft)

        parsed = json.loads(json_buf.getvalue().decode("utf-8"))
        self.assertEqual(parsed["draft"]["offer_id"], draft.offer_id)
        self.assertGreater(len(xlsx_buf.getvalue()), 1000)
        with ZipFile(xlsx_buf) as archive:
            self.assertIn("xl/workbook.xml", archive.namelist())

    def test_preview_explains_next_steps_for_draft(self):
        draft = build_ozon_card_draft("товар: Кусачки маникюрные\nцена: 444\nhttps://example.com/photo.jpg")
        preview = format_ozon_card_preview(draft)

        self.assertIn("нужна ручная доработка перед загрузкой", preview)
        self.assertIn("категория", preview)
        self.assertIn("Откройте XLSX", preview)
        self.assertIn("JSON пока нужен как техфайл", preview)

    def test_batch_exports_xlsx_and_json(self):
        ready = build_ozon_card_draft(
            "товар: Кабель USB-C\nкатегория: Кабели\nцена: 299\nвес: 50 г\nразмер: 100x80x20 мм\nhttps://example.com/photo.jpg"
        )
        needs_review = build_ozon_card_draft("товар: Кусачки маникюрные\nцена: 444")
        items = [
            OzonCardBatchItem(ready, "ready source", "ready"),
            OzonCardBatchItem(needs_review, "review source", "needs_review", "категория, фото"),
            OzonCardBatchItem(None, "bad source", "error", "parse error"),
        ]

        json_buf = export_ozon_cards_batch_json(items)
        parsed = json.loads(json_buf.getvalue().decode("utf-8"))
        self.assertEqual(parsed["summary"]["total"], 3)
        self.assertEqual(parsed["summary"]["ready"], 1)
        self.assertEqual(parsed["summary"]["needs_review"], 1)
        self.assertEqual(parsed["summary"]["error"], 1)

        xlsx_buf = export_ozon_cards_batch_xlsx(items)
        self.assertGreater(len(xlsx_buf.getvalue()), 1000)
        with ZipFile(xlsx_buf) as archive:
            self.assertIn("xl/workbook.xml", archive.namelist())

    def test_batch_preview_shows_summary_and_problem_items(self):
        draft = build_ozon_card_draft("товар: Кусачки маникюрные\nцена: 444")
        preview = format_ozon_batch_preview([OzonCardBatchItem(draft, "source", "needs_review")])

        self.assertIn("Пакет карточек Ozon готов", preview)
        self.assertIn("Нужна доработка", preview)
        self.assertIn("категория", preview)

    async def test_enhanced_card_uses_ai_json_when_available(self):
        ai_response = json.dumps({
            "name": "Кабель USB-C для зарядки и передачи данных",
            "description": "Надежный кабель USB-C для ежедневной зарядки устройств и передачи данных.",
            "category_hint": "Кабели",
            "brand": "Нет бренда",
            "attributes": {"Назначение": "зарядка и передача данных"},
            "keywords": ["кабель usb-c", "кабель для зарядки"],
            "checklist": ["Проверить длину кабеля"],
            "notes": "AI улучшил название и описание без добавления неподтвержденных фактов.",
        })
        with (
            patch("app.card_filler.ai_is_available", return_value=True),
            patch("app.card_filler.ask_ai", new=AsyncMock(return_value=ai_response)),
        ):
            draft = await build_enhanced_ozon_card_draft("товар: Кабель USB-C\nцена: 299")

        self.assertEqual(draft.name, "Кабель USB-C для зарядки и передачи данных")
        self.assertIn("ежедневной зарядки", draft.description)
        self.assertEqual(draft.attributes["Назначение"], "зарядка и передача данных")
        self.assertEqual(draft.checklist, ["Проверить длину кабеля"])

    async def test_enhanced_card_uses_competitor_context_without_ai(self):
        with patch("app.card_filler.ai_is_available", return_value=False):
            draft = await build_enhanced_ozon_card_draft(
                "товар: Держатель телефона\nцена: 599",
                competitors=[
                    {"name": "Держатель телефона автомобильный магнитный", "price": 500},
                    {"name": "Держатель телефона в машину на панель", "price": 700},
                    {"name": "Автомобильный держатель смартфона", "price": 900},
                ],
            )

        self.assertIn("Найдено конкурентов: 3", draft.competitor_summary)
        self.assertIn("медиана 700", draft.competitor_summary)
        self.assertIn("держатель", draft.keywords)
        self.assertIn("Проверить, чем карточка отличается от конкурентов в выдаче Ozon", draft.checklist)

    async def test_enhanced_card_handles_ai_error_as_local_draft(self):
        with (
            patch("app.card_filler.ai_is_available", return_value=True),
            patch("app.card_filler.ask_ai", new=AsyncMock(return_value="AI error: network")),
        ):
            draft = await build_enhanced_ozon_card_draft("товар: Кабель USB-C\nцена: 299")

        self.assertIn("AI-доработка недоступна", draft.ai_notes)
        self.assertIn("Кабель USB-C", draft.name)


    def test_keywords_cleanup_drops_technical_tokens(self):
        draft = build_ozon_card_draft(
            "товар: Коврик для йоги\n"
            "бренд: YOGAFOX\n"
            "описание: https://images.wb.ru/vol1/basket/item.webp catalog wildberries"
        )
        lowered = {k.lower() for k in draft.keywords}
        self.assertNotIn("basket", lowered)
        self.assertNotIn("vol", lowered)
        self.assertNotIn("webp", lowered)
        self.assertNotIn("images", lowered)
        self.assertNotIn("wildberries", lowered)

    def test_category_mapping_wb_to_ozon_path(self):
        draft = build_ozon_card_draft("категория wb: Мебель\nтовар: кресло офисное")
        self.assertEqual(draft.category_hint, "Дом и сад/Мебель/Кресла")

    def test_xlsx_starts_from_header_without_service_rows(self):
        draft = build_ozon_card_draft("товар: Коврик для йоги")
        xlsx_buf = export_ozon_card_xlsx(draft)
        wb = openpyxl.load_workbook(xlsx_buf)
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Поле")
        self.assertEqual(ws["B1"].value, "Значение")

    def test_keywords_respect_category_whitelist(self):
        draft = build_ozon_card_draft(
            "категория wb: Спорт и отдых/Фитнес и йога/Коврики для йоги\n"
            "товар: Коврик для йоги нескользящий\n"
            "описание: стильный подарок для дома"
        )
        lowered = {k.lower() for k in draft.keywords}
        self.assertIn("коврик", lowered)
        self.assertNotIn("подарок", lowered)

    def test_profile_filters_forbidden_words_and_limits_description(self):
        draft = build_ozon_card_draft("товар: лучший коврик для йоги акция\nописание: лучший выбор для дома")
        apply_card_profile(
            draft,
            {
                "forbidden_words": ["лучший", "акция"],
                "max_length": 20,
                "required_attributes": ["Материал"],
            },
        )
        self.assertNotIn("лучший", draft.name.lower())
        self.assertNotIn("акция", draft.name.lower())
        self.assertLessEqual(len(draft.description), 20)
        self.assertIn("Материал", draft.attributes)


if __name__ == "__main__":
    unittest.main()

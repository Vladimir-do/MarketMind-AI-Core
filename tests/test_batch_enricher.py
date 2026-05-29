import tempfile
import unittest
from pathlib import Path

import openpyxl

from app.batch_enricher import enrich_file


class BatchEnricherTests(unittest.TestCase):
    def test_enrich_xlsx_adds_agent_columns_and_checkpoint(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["id", "товар", "цена", "вес", "размер"])
            ws.append([1, "Держатель телефона автомобильный", "599 руб", "180 г", "120x80x70 мм"])
            ws.append([2, "Кабель USB-C", "299 руб", "50 г", "100x80x20 мм"])
            wb.save(input_path)

            result = enrich_file(input_path, output_path, limit=1)

            self.assertEqual(result["total_rows"], 2)
            self.assertEqual(result["processed"], 1)
            self.assertTrue(output_path.exists())
            self.assertTrue(output_path.with_suffix(".xlsx.checkpoint.json").exists())

            out_wb = openpyxl.load_workbook(output_path)
            out_ws = out_wb.active
            headers = [cell.value for cell in out_ws[1]]
            self.assertIn("agent_name", headers)
            self.assertIn("agent_ozon_payload_json", headers)

    def test_resume_skips_processed_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "input.xlsx"
            output_path = Path(tmp) / "output.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["id", "товар", "цена"])
            ws.append([1, "Держатель телефона", "599 руб"])
            wb.save(input_path)

            enrich_file(input_path, output_path, limit=1)
            result = enrich_file(input_path, output_path, resume=True)

            self.assertEqual(result["processed"], 0)
            self.assertEqual(result["skipped"], 1)


if __name__ == "__main__":
    unittest.main()

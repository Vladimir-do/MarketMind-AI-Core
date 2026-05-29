import tempfile
import unittest
from pathlib import Path
from zipfile import ZipFile

import openpyxl
from unittest.mock import patch

from app.brd_enricher import (
    BRD_REQUIRED_COLUMNS,
    BrdSource,
    brady_search_query,
    classify_brd_category,
    load_docx_lines,
    prepare_brd_table,
    research_brd_article,
)


def make_docx(path: Path, lines: list[str]) -> None:
    body = "".join(
        f"<w:p><w:r><w:t>{line}</w:t></w:r></w:p>"
        for line in lines
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{body}</w:body>"
        "</w:document>"
    )
    with ZipFile(path, "w") as archive:
        archive.writestr("word/document.xml", document)


class BrdEnricherTests(unittest.TestCase):
    def test_brd_search_query_strips_prefix(self):
        self.assertEqual(brady_search_query("brd622695"), "Brady 622695")

    def test_load_docx_lines(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "Категории BRD.docx"
            make_docx(path, ["Термотрансферные принтеры", "Прочее"])
            self.assertEqual(load_docx_lines(path), ["Термотрансферные принтеры", "Прочее"])

    def test_prepare_brd_table_adds_service_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "Таблица BRD.xlsx"
            output_path = tmp_path / "Таблица BRD ИИ.xlsx"
            categories_path = tmp_path / "Категории BRD.docx"
            make_docx(categories_path, ["Термотрансферные принтеры", "Прочее"])

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BRD ТбД"
            ws.append(BRD_REQUIRED_COLUMNS)
            ws.append(["brd101954", None, None, None, None, None, None])
            wb.save(input_path)

            result = prepare_brd_table(
                input_path,
                output_path,
                categories_path=categories_path,
                limit=1,
            )

            self.assertEqual(result["total_rows"], 1)
            self.assertEqual(result["processed"], 1)
            self.assertTrue(output_path.exists())
            out_wb = openpyxl.load_workbook(output_path)
            out_ws = out_wb["BRD ТбД"]
            headers = [cell.value for cell in out_ws[1]]
            self.assertIn("ИИ поисковый запрос", headers)
            query_col = headers.index("ИИ поисковый запрос") + 1
            status_col = headers.index("ИИ статус") + 1
            self.assertEqual(out_ws.cell(2, query_col).value, "Brady 101954")
            self.assertEqual(out_ws.cell(2, status_col).value, "needs_online_research")

    def test_research_brd_article_fills_from_verified_sources(self):
        categories = ["Термотрансферные принтеры", "Материалы для портативных термотрансферных принтеров", "Прочее"]
        source = BrdSource(
            url="https://example.com/brady-101954",
            title="Brady 101954 Portable Printer Label Cartridge",
            text=(
                "Brady 101954 Portable Printer Label Cartridge. "
                "The 101954 cartridge is compatible with BMP21 printers.\n"
                "Material: vinyl\nColor: white\nWidth: 19 mm"
            ),
            image_url="https://example.com/image.jpg",
        )

        with (
            patch("app.brd_enricher.search_web_urls", return_value=[source.url]),
            patch("app.brd_enricher.fetch_brd_source", return_value=source),
            patch("app.brd_enricher._download_first_image", return_value="brd101954.jpg"),
        ):
            result = research_brd_article("brd101954", categories, img_dir=Path("img"))

        self.assertEqual(result.status, "filled_online")
        self.assertIn("Brady 101954", result.name)
        self.assertIn("<p>", result.description_html)
        self.assertIn("<ul>", result.specs_html)
        self.assertEqual(result.image_filename, "brd101954.jpg")
        self.assertEqual(result.sources, [source.url])
        self.assertGreater(result.confidence, 0.5)

    def test_arc_flash_label_category_wins_over_spills_copy(self):
        safety_signs = "\u0413\u043e\u0442\u043e\u0432\u044b\u0435 \u0437\u043d\u0430\u043a\u0438 \u0431\u0435\u0437\u043e\u043f\u0430\u0441\u043d\u043e\u0441\u0442\u0438"
        sorbents = "\u0421\u043e\u0440\u0431\u0435\u043d\u0442\u044b \u0434\u043b\u044f \u043f\u0440\u0435\u0434\u043e\u0442\u0432\u0440\u0430\u0449\u0435\u043d\u0438\u044f \u0438 \u043b\u0438\u043a\u0432\u0438\u0434\u0430\u0446\u0438\u044f \u043f\u0440\u043e\u043b\u0438\u0432\u043e\u0432"
        text = "Arc Flash Protection Label protected graphics hold up to repeated spills and cleanings"

        self.assertEqual(classify_brd_category(text, [sorbents, safety_signs]), safety_signs)

    def test_prepare_brd_table_online_writes_product_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            input_path = tmp_path / "Таблица BRD.xlsx"
            output_path = tmp_path / "Таблица BRD ИИ.xlsx"
            categories_path = tmp_path / "Категории BRD.docx"
            make_docx(categories_path, ["Материалы для портативных термотрансферных принтеров", "Прочее"])

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BRD ТбД"
            ws.append(BRD_REQUIRED_COLUMNS)
            ws.append(["brd101954", None, None, None, None, None, None])
            wb.save(input_path)

            source = BrdSource(
                url="https://example.com/brady-101954",
                title="Brady 101954 Portable Printer Label Cartridge",
                text="Brady 101954 cartridge. 101954 compatible with BMP21. Material: vinyl",
                image_url="https://example.com/image.jpg",
            )
            with (
                patch("app.brd_enricher.search_web_urls", return_value=[source.url]),
                patch("app.brd_enricher.fetch_brd_source", return_value=source),
                patch("app.brd_enricher._download_first_image", return_value="brd101954.jpg"),
            ):
                prepare_brd_table(
                    input_path,
                    output_path,
                    categories_path=categories_path,
                    limit=1,
                    online=True,
                    delay_sec=0,
                )

            out_wb = openpyxl.load_workbook(output_path)
            out_ws = out_wb["BRD ТбД"]
            headers = [cell.value for cell in out_ws[1]]
            row = {header: out_ws.cell(2, idx + 1).value for idx, header in enumerate(headers)}
            self.assertIn("Brady 101954", row["Наименование"])
            self.assertIn("<p>", row["Описание"])
            self.assertIn("<ul>", row["Характеристики"])
            self.assertEqual(row["Картинки"], "brd101954.jpg")
            self.assertEqual(row["ИИ статус"], "filled_online")


if __name__ == "__main__":
    unittest.main()
